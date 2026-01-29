import win32com.client
import pandas as pd
import shutil
import pdfplumber
from pathlib import Path
from io import StringIO
import logging
import re
from datetime import datetime
import sys
import warnings
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl import Workbook

# Suprimir avisos
warnings.filterwarnings('ignore')
2
# --- 1. CONFIGURAÇÕES GERAIS ---
CAMINHO_PASTA_LOCAL = Path(r"C:\Users\matheus.augusto\OneDrive - Grupo Fleury\Planejamento Financeiro - TI&Telecom _ - Documentos\RELATÓRIO\Automações\Gerar Relatório OPEX")

# Data de Corte (Ignora e-mails antigos)
DATA_INICIO_LEITURA = datetime(2026, 1, 1)

if not Path.exists(CAMINHO_PASTA_LOCAL):
    try:
        CAMINHO_PASTA_LOCAL.mkdir(parents=True, exist_ok=True)
        logging.info(f"Pasta criada em: {CAMINHO_PASTA_LOCAL}")
        print(f"Pasta criada em: {CAMINHO_PASTA_LOCAL}")
    except:
        pass

ARQUIVO_FINAL = CAMINHO_PASTA_LOCAL / "Relatorios_OPEX.xlsx"
PASTA_LOGS = CAMINHO_PASTA_LOCAL / "Logs"
PASTA_LOGS.mkdir(exist_ok=True)

# Configuração de Log
logging.basicConfig(
    filename=PASTA_LOGS / f"log_execucao_{datetime.now().strftime('%Y-%m-%d_%H-%M-%S')}.txt",
    filemode='w', 
    level=logging.INFO,
    format='%(asctime)s - %(levelname)s - %(message)s',
    datefmt='%d/%m/%Y %H:%M:%S'
)

# --- LISTA NEGRA DE REMETENTES ---
REMETENTES_IGNORAR = [
    "gleiciane.fragoso@grupofleury.com.br",
]

# Mapa de Meses
MAPA_MESES_INFO = {
    'january': ('Janeiro', '01'), 'jan': ('Janeiro', '01'), 'janeiro': ('Janeiro', '01'), '01': ('Janeiro', '01'),
    'february': ('Fevereiro', '02'), 'feb': ('Fevereiro', '02'), 'fevereiro': ('Fevereiro', '02'), '02': ('Fevereiro', '02'),
    'march': ('Março', '03'), 'mar': ('Março', '03'), 'março': ('Março', '03'), '03': ('Março', '03'),
    'april': ('Abril', '04'), 'apr': ('Abril', '04'), 'abril': ('Abril', '04'), '04': ('Abril', '04'),
    'may': ('Maio', '05'), 'maio': ('Maio', '05'), '05': ('Maio', '05'),
    'june': ('Junho', '06'), 'jun': ('Junho', '06'), 'junho': ('Junho', '06'), '06': ('Junho', '06'),
    'july': ('Julho', '07'), 'jul': ('Julho', '07'), 'julho': ('Julho', '07'), '07': ('Julho', '07'),
    'august': ('Agosto', '08'), 'aug': ('Agosto', '08'), 'agosto': ('Agosto', '08'), '08': ('Agosto', '08'),
    'september': ('Setembro', '09'), 'sep': ('Setembro', '09'), 'setembro': ('Setembro', '09'), 'set': ('Setembro', '09'), '09': ('Setembro', '09'),
    'october': ('Outubro', '10'), 'oct': ('Outubro', '10'), 'outubro': ('Outubro', '10'), 'out': ('Outubro', '10'), '10': ('Outubro', '10'),
    'november': ('Novembro', '11'), 'nov': ('Novembro', '11'), 'novembro': ('Novembro', '11'), '11': ('Novembro', '11'),
    'december': ('Dezembro', '12'), 'dec': ('Dezembro', '12'), 'dezembro': ('Dezembro', '12'), 'dez': ('Dezembro', '12'), '12': ('Dezembro', '12')
}

# Colunas padrão para garantir ordem
COLUNAS_PADRAO = [
    'Mes_Referencia', 'Data_Email', 'Remetente', 
    'Assunto_Email', 'Categoria_OPEX', 'Data_Processamento'
] 

# --- CONFIGURAÇÃO DOS FORNECEDORES ---
CONFIG_PADRAO = [
    {
        "Fornecedor": "Selbetti",
        "Palavras_Chave": "Faturamento Selbetti, Relatório Selbetti, RE: Faturamento Selbetti",
        "Nome_Aba": "Selbetti",
        "Categoria_OPEX": "Impressoras/Impressão",
        "Tipo_leitura": "Corpo"
    },
    {
        "Fornecedor": "Daycoval",
        "Palavras_Chave": "Faturamento Banco Daycoval, Relatório Daycoval",
        "Nome_Aba": "Daycoval",
        "Categoria_OPEX": "DAYCOVAL LEASING TI",
        "Tipo_leitura": "Corpo"
    },
    {
        "Fornecedor": "Positivo",
        "Palavras_Chave": "Faturamento Positivo, Locação",
        "Nome_Aba": "Positivo",
        "Categoria_OPEX": "POSITIVO LEASING TI",
        "Tipo_leitura": "Corpo"
    },
        {
        "Fornecedor": "PDF",
        "Palavras_Chave": "PDF",
        "Nome_Aba": "PDF",
        "Categoria_OPEX": "PDF LEASING TI",
        "Tipo_leitura": "PDF"
    },    {
        "Fornecedor": "Ambos",
        "Palavras_Chave": "Ambos",
        "Nome_Aba": "Ambos",
        "Categoria_OPEX": "Ambos LEASING TI",
        "Tipo_leitura": "Ambos"
    }, {
        "Fornecedor": "TESTE",
        "Palavras_Chave": "TESTE, TESTE1, TESTE",
        "Nome_Aba": "TESTE",
        "Categoria_OPEX": "TESTE LEASING TI",
        "Tipo_leitura": "Corpo"
    }, {
        "Fornecedor": "NOVA",
        "Palavras_Chave": "NOVA",
        "Nome_Aba": "NOVA",
        "Categoria_OPEX": "NOVA LEASING TI",
        "Tipo_leitura": "Corpo"
    }
]

# --- FUNÇÕES AUXILIARES ---

def aplicar_estilo_visual(ws, sheet_name):
    """
    Aplica formatação visual e corrige a definição da Tabela Oficial para evitar corrupção.
    """
    max_col = ws.max_column
    max_row = ws.max_row
    letra_ultima_coluna = get_column_letter(max_col)
    
    # --- DEFINIÇÃO DE CORES ---
    # Azul Escuro: Colunas Geradas pelo Robô (Metadados)
    fill_meta = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") 
    # Laranja: Colunas Extraídas do E-mail/PDF (Dados Originais)
    fill_dados = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")
    
    font_branca = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
    
    # Lista das colunas que o robô cria (Metadados)
    cols_meta_nomes = ['Mes_Referencia', 'Data_Email', 'Remetente', 'Assunto_Email', 
                       'Categoria_OPEX', 'Data_Processamento', 'Chave_Negocio_Temp']

    # 1. Formata o Cabeçalho (Linha 1)
    for cell in ws[1]:
        col_name = str(cell.value)
        cell.font = font_branca
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Se o nome da coluna estiver na lista de metadados, pinta de Azul
        if col_name in cols_meta_nomes:
            cell.fill = fill_meta
        else:
            # Caso contrário (dado extraído da tabela original), pinta de Laranja
            cell.fill = fill_dados

    # 2. Ajuste de Largura das Colunas
    for i in range(1, max_col + 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 20

    # 3. Criação/Correção da Tabela Oficial
    nome_tabela_limpo = f"TB_{sheet_name.replace(' ', '_')}"
    ref_tabela = f"A1:{letra_ultima_coluna}{max_row}"
    
    # Verifica se a tabela já existe na planilha
    tabela_existente = None
    if ws.tables:
        for tbl in ws.tables.values():
            if tbl.name == nome_tabela_limpo:
                tabela_existente = tbl
                break
    
    if tabela_existente:
        # Se existe, APENAS atualiza o range (não recria)
        tabela_existente.ref = ref_tabela
    else:
        # Se não existe, cria do zero
        tabela = Table(displayName=nome_tabela_limpo, ref=ref_tabela)
        # Estilo 'TableStyleLight1' deixa as linhas de dados brancas/limpas
        estilo = TableStyleInfo(name="TableStyleLight1", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
        tabela.tableStyleInfo = estilo
        ws.add_table(tabela)


def extrair_tabelas_de_pdf(caminho_pdf):
    """
    Tenta extrair tabelas de um PDF usando múltiplas estratégias (Linhas e Fluxo).
    Versão robusta para lidar com diferentes formatos de PDF.
    """
    lista_dfs = []
    try:
        with pdfplumber.open(caminho_pdf) as pdf:
            for page in pdf.pages:
                # TENTATIVA 1: Lattice (Tabelas com bordas/linhas explícitas)
                tabelas = page.extract_tables({"vertical_strategy": "lines", "horizontal_strategy": "lines"})
                
                # TENTATIVA 2: Stream (Tabelas baseadas em espaçamento de texto - comum em notas fiscais)
                if not tabelas:
                    tabelas = page.extract_tables({"vertical_strategy": "text", "horizontal_strategy": "text"})
                
                # TENTATIVA 3: Padrão (Deixa o pdfplumber tentar adivinhar)
                if not tabelas:
                    tabelas = page.extract_tables()

                for tabela in tabelas:
                    # Limpeza: Remove linhas que estão totalmente vazias ou só contêm None/espaços
                    tabela_limpa = [
                        [str(celula).replace('\n', ' ').strip() if celula is not None else '' for celula in linha]
                        for linha in tabela
                        if any(celula is not None and str(celula).strip() != "" for celula in linha)
                    ]
                    
                    # Precisa ter pelo menos cabeçalho + 1 linha de dados
                    if len(tabela_limpa) > 1:
                        # Cria DataFrame assumindo que a 1ª linha é o cabeçalho
                        # (O tratamento de cabeçalho errado será feito depois na 'encontrar_cabecalho_correto')
                        df = pd.DataFrame(tabela_limpa[1:], columns=tabela_limpa[0])
                        lista_dfs.append(df)
                        
    except Exception as e:
        logging.error(f"Erro ao ler PDF {caminho_pdf}: {e}")
        print(f"   [ERRO PDF] Falha na leitura: {e}")
        
    return lista_dfs

def is_file_open(path):
    if not path.exists(): return False
    try:
        os.rename(path, path)
        return False
    except OSError:
        return True

def eh_coluna_financeira(nome_coluna):
    """
    Central de Controle: Define quais colunas são dinheiro.
    """
    nome_coluna = str(nome_coluna).lower()
    TERMOS_MOEDA = [
        'valor', 'total', 'liquido', 'bruto', 'realizado', 
        'vlr', 'custo', 'taxa', 'imposto', 'montante', 'r$'
    ]
    return any(termo in nome_coluna for termo in TERMOS_MOEDA)

def obter_email_remetente(msg):
    try:
        sender = msg.SenderEmailAddress
        if not sender: return ""
        if "O=" in sender and "@" not in sender:
            try:
                property_accessor = msg.Sender.PropertyAccessor
                sender = property_accessor.GetProperty("http://schemas.microsoft.com/mapi/proptag/0x39FE001E")
            except: pass
        return sender.lower()
    except: return ""

def extrair_info_data_inteligente(assunto, corpo):
    padrao = r'(janeiro|jan|01|fevereiro|feb|02|março|mar|03|abril|apr|04|maio|may|05|junho|jun|06|julho|jul|07|agosto|aug|08|setembro|sep|set|09|outubro|oct|out|10|novembro|nov|11|dezembro|dec|dez|12)(?:/?\s?(?:de)?\s?(\d{2,4}))?'
    match = re.search(padrao, assunto.lower())
    if not match: match = re.search(padrao, corpo.lower())
    ano_atual = str(datetime.now().year)
    if match:
        mes_raw = match.group(1)
        ano_raw = match.group(2)
        mes_pt, mes_num = MAPA_MESES_INFO.get(mes_raw, (mes_raw.capitalize(), '00'))
        if ano_raw:
            if len(ano_raw) == 4 and ano_raw.startswith("20"): ano_final = ano_raw
            elif len(ano_raw) == 2: ano_final = "20" + ano_raw
            else: ano_final = ano_atual
        else: ano_final = ano_atual
        return mes_pt, mes_num, ano_final
    return None, None, ano_atual

def extrair_data_da_tabela(df):
    termos_data = ['data emissão', 'data emissao', 'dt emissao', 'data frs', 'dt frs', 'referencia', 'competencia', 'data de emissão', 'emissao']
    coluna_data_encontrada = None
    for col in df.columns:
        col_str = str(col).lower()
        if any(termo in col_str for termo in termos_data):
            coluna_data_encontrada = col
            break
    if not coluna_data_encontrada: return None, None, None
    try:
        valores = df[coluna_data_encontrada].dropna()
        if valores.empty: return None, None, None
        valor_data = valores.iloc[0]
        if isinstance(valor_data, str):
            try: dt = pd.to_datetime(valor_data, dayfirst=True)
            except: return extrair_info_data_inteligente(valor_data, "")
        elif isinstance(valor_data, (datetime, pd.Timestamp)): dt = valor_data
        else: return None, None, None
        
        meses_lista = ['Indef', 'Janeiro', 'Fevereiro', 'Março', 'Abril', 'Maio', 'Junho', 
                       'Julho', 'Agosto', 'Setembro', 'Outubro', 'Novembro', 'Dezembro']
        mes_pt = meses_lista[dt.month] if 1 <= dt.month <= 12 else "Indefinido"
        return mes_pt, f"{dt.month:02d}", str(dt.year)
    except Exception as e:
        return None, None, None

def limpar_valor_monetario(valor):
    if isinstance(valor, (int, float)): return valor
    if isinstance(valor, str):
        limpo = valor.replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
        try: return float(limpo)
        except: return 0.0
    return 0.0

def encontrar_cabecalho_correto(df, colunas_esperadas):
    """
    Procura a linha de cabeçalho real, tolerando a sujeira comum em PDFs.
    """
    # Termos-chave que indicam fortemente uma tabela financeira
    termos_validos = ['valor', 'total', 'liquido', 'bruto', 'vlr', 'data', 'emissao', 'nota', 'nf', 'r$', 'montante', 'descrição', 'historico']
    
    # Função interna para verificar se uma lista de colunas (linha) parece ser um cabeçalho válido
    def eh_cabecalho_valido(lista_colunas):
        cols_str = [str(c).lower().strip() for c in lista_colunas]
        # Retorna True se encontrar pelo menos UM dos termos válidos na linha
        return any(t in c for c in cols_str for t in termos_validos)

    # 1. Tenta o cabeçalho atual do DataFrame
    if eh_cabecalho_valido(df.columns):
        return df

    # 2. Varre as primeiras 15 linhas (PDFs costumam ter cabeçalhos grandes, logotipos ou textos introdutórios antes da tabela real)
    for i in range(min(15, len(df))):
        linha = df.iloc[i]
        
        if eh_cabecalho_valido(linha):
            # Promove esta linha a cabeçalho
            df_novo = df[i+1:].copy() # Pega os dados abaixo da linha encontrada
            df_novo.columns = [str(x).replace('\n', ' ').strip() for x in linha] # Define a linha como novo cabeçalho
            
            # Reset do index para manter a integridade
            df_novo.reset_index(drop=True, inplace=True)
            return df_novo
            
    return None

def tratar_dataframe(df, config, metadados):
    if "colunas_renomear" in config: df = df.rename(columns=config["colunas_renomear"])
    df['Data_Email'] = pd.to_datetime(metadados['data_recebimento'], utc=True).tz_localize(None)
    df['Remetente'] = metadados['remetente']
    df['Assunto_Email'] = metadados['assunto']
    df['Mes_Referencia'] = metadados['mes_nome_pt'] if metadados['mes_nome_pt'] else ""
    df['Categoria_OPEX'] = config['classificacao_opex']
    df['Data_Processamento'] = datetime.now()

    # Aplica limpeza apenas em colunas financeiras
    colunas_valor = [col for col in df.columns if eh_coluna_financeira(col)]
    for col in colunas_valor: 
        df[col] = df[col].apply(limpar_valor_monetario)

    mes_chave = metadados['mes_nome_pt'] if metadados['mes_nome_pt'] else "SEM_DATA"
    col_desc = next((c for c in df.columns if 'Descricao' in str(c) or 'Linha' in str(c)), df.columns[0])
    
    # Tenta achar coluna de valor para a chave, senão pega a segunda coluna
    col_val = df.columns[1] if len(df.columns) > 1 else df.columns[0]
    for c in df.columns:
        if eh_coluna_financeira(c):
            col_val = c
            break
    
    df['Chave_Negocio_Temp'] = mes_chave + metadados['ano_full'] + "_" + \
                               df[col_desc].astype(str) + "_" + \
                               df[col_val].astype(str)
    
    cols_dados = [c for c in df.columns if c not in COLUNAS_PADRAO and c != 'Chave_Negocio_Temp']
    cols_finais = ['Mes_Referencia'] + cols_dados + ['Data_Email', 'Remetente', 'Assunto_Email', 'Categoria_OPEX', 'Data_Processamento', 'Chave_Negocio_Temp']
    
    return df[cols_finais]

def salvar_com_append_preservando_formatacao(df_novos, caminho_arquivo, nome_aba):
    if df_novos.empty: return

    # Remove a coluna temporária do DataFrame antes de salvar (para não ir pro Excel)
    df_dados = df_novos.copy()
    if 'Chave_Negocio_Temp' in df_dados.columns:
        df_dados = df_dados.drop(columns=['Chave_Negocio_Temp'])

    # --- CENÁRIO 1: ARQUIVO NOVO ---
    if not caminho_arquivo.exists():
        with pd.ExcelWriter(caminho_arquivo, engine='openpyxl') as writer:
            df_dados.to_excel(writer, sheet_name=nome_aba, index=False)
            # Aplica o estilo na aba recém-criada
            aplicar_estilo_visual(writer.sheets[nome_aba], nome_aba)
        return

    try:
        # Carrega o workbook existente
        wb = load_workbook(caminho_arquivo)
        
        # --- CENÁRIO 2: ABA NOVA ---
        if nome_aba not in wb.sheetnames:
            ws = wb.create_sheet(nome_aba)
            # Escreve cabeçalho
            ws.append(list(df_dados.columns))
            # Escreve dados
            for r in dataframe_to_rows(df_dados, index=False, header=False):
                ws.append(r)
            
            aplicar_estilo_visual(ws, nome_aba)
            wb.save(caminho_arquivo)
            return

        # --- CENÁRIO 3: APPEND EM ABA EXISTENTE ---
        ws = wb[nome_aba]
        
        # Lê cabeçalho existente para garantir a ordem das colunas
        headers_excel = [cell.value for cell in ws[1]]
        
        # Alinha as colunas do DF com as do Excel
        for col in headers_excel:
            if col not in df_dados.columns:
                df_dados[col] = "" # Cria coluna vazia se faltar
        
        # Reordena o DF para bater com o Excel
        df_dados = df_dados[headers_excel]

        # Escreve apenas as linhas novas
        for r in dataframe_to_rows(df_dados, index=False, header=False):
            ws.append(r)
        
        # --- PASSO CRUCIAL PARA NÃO CORROMPER ---
        # Chama a função de estilo para atualizar o range da tabela existente
        aplicar_estilo_visual(ws, nome_aba)
        
        wb.save(caminho_arquivo)
        print(f"   -> Dados salvos em '{nome_aba}' com sucesso.")

    except Exception as e:
        logging.error(f"Erro ao salvar em {nome_aba}: {e}")
        print(f"   -> ERRO AO SALVAR EXCEL: {e}")

def aplicar_estilo_inicial(ws, sheet_name):
    """
    Aplica estilo e CRIA a tabela oficial pela primeira vez nas abas de dados.
    """
    max_col = ws.max_column
    letra_ultima_coluna = get_column_letter(max_col)
    max_row = ws.max_row
    
    # Estilo do Cabeçalho
    fill_azul = PatternFill(start_color="2F75B5", end_color="2F75B5", fill_type="solid")
    font_branca = Font(name="Calibri", size=11, color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = fill_azul
        cell.font = font_branca
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
    # Ajuste de Colunas
    ws.column_dimensions['A'].width = 15 # Mes
    ws.column_dimensions['B'].width = 18 # Data
    ws.column_dimensions['C'].width = 25 # Remetente
    ws.column_dimensions['D'].width = 35 # Assunto
    
    # Criação da Tabela Oficial
    nome_tabela = f"TB_{sheet_name.replace(' ', '_')}"
    ref_tabela = f"A1:{letra_ultima_coluna}{max_row}"
    
    tabela = Table(displayName=nome_tabela, ref=ref_tabela)
    estilo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    tabela.tableStyleInfo = estilo
    
    # Só adiciona se não existir
    if nome_tabela not in ws.tables:
        ws.add_table(tabela)

# --- PROCESSO PRINCIPAL ---

def realizar_backup_seguranca():
    """
    Cria uma cópia do arquivo Excel atual antes de qualquer alteração.
    O arquivo é salvo na pasta 'Backups' com data e hora no nome.
    """
    if not ARQUIVO_FINAL.exists():
        logging.info("Arquivo final não existe. Pulinho backup.")
        print("Arquivo final não existe. Pulinho backup.")
        return

    try:
        # 1. Define onde será a pasta de backups
        pasta_backup = CAMINHO_PASTA_LOCAL / "Backups"
        
        # 2. Cria a pasta se ela não existir
        pasta_backup.mkdir(exist_ok=True)
        
        # 3. Define o nome do arquivo com DATA e HORA (para não substituir o anterior)
        timestamp = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
        nome_backup = f"Backup_{timestamp}_Relatorios_OPEX.xlsx"
        caminho_backup = pasta_backup / nome_backup
        
        # 4. Copia o arquivo
        shutil.copy2(ARQUIVO_FINAL, caminho_backup)
        
        print(f" -> BACKUP REALIZADO: {nome_backup}")
        logging.info(f"Backup de segurança criado: {nome_backup}")
        
    except Exception as e:
        print(f"ERRO AO FAZER BACKUP: {e}")
        logging.error(f"Falha no backup: {e}")    

def inicializar_aba_config():
    """
    Cria a aba 'config_fornecedor' com Tabela Oficial TB_ConfigFornecedor.
    """
    try:
        wb = None
        salvar_necessario = False
        NOME_ABA_CONFIG = "config_fornecedor"

        if not ARQUIVO_FINAL.exists():
            print(f"--- CRIANDO ARQUIVO NOVO EM: {ARQUIVO_FINAL} ---")
            wb = Workbook() 
            if 'Sheet' in wb.sheetnames: del wb['Sheet']
            salvar_necessario = True
        else:
            wb = load_workbook(ARQUIVO_FINAL)

        if NOME_ABA_CONFIG not in wb.sheetnames:
            print(f"Criando aba '{NOME_ABA_CONFIG}'...")
            ws = wb.create_sheet(NOME_ABA_CONFIG, 0)

            cabecalho = ["Fornecedor", "Palavras_Chave", "Nome_Aba", "Categoria_OPEX", "Tipo_leitura"]
            ws.append(cabecalho)

            dados = CONFIG_PADRAO
            if isinstance(CONFIG_PADRAO, dict): dados = list(CONFIG_PADRAO.values())

            for item in dados:
                if isinstance(item, dict):
                    ws.append([
                        item.get("Fornecedor", ""),
                        item.get("Palavras_Chave", ""),
                        item.get("Nome_Aba", ""),
                        item.get("Categoria_OPEX", ""),
                        item.get("Tipo_leitura", "Corpo")
                    ])

            # --- ESTILO ---
            # Formata APENAS o cabeçalho
            fill_header = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # Azul Escuro
            font_header = Font(bold=True, color="FFFFFF")
            
            for cell in ws[1]:
                cell.fill = fill_header
                cell.font = font_header
                cell.alignment = Alignment(horizontal='center')

            # Ajuste de largura
            ws.column_dimensions['A'].width = 15
            ws.column_dimensions['B'].width = 40
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 25
            ws.column_dimensions['E'].width = 15

            # --- CRIAÇÃO DA TABELA OFICIAL ---
            nome_tabela = "TB_ConfigFornecedor"
            ref_tabela = f"A1:E{ws.max_row}"
            tabela = Table(displayName=nome_tabela, ref=ref_tabela)
            
            # Estilo 'TableStyleLight8' ou 'None' é bem limpo (fundo branco, cabeçalho sutil)
            # Se quiser algo padrão do Excel, use 'TableStyleMedium2'
            estilo = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)
            tabela.tableStyleInfo = estilo
            ws.add_table(tabela)

            # --- LISTA SUSPENSA ---
            dv = DataValidation(type="list", formula1='"Corpo,PDF,Ambos"', allow_blank=True)
            ws.add_data_validation(dv)
            dv.add(f'E2:E{ws.max_row + 50}') 
            
            salvar_necessario = True

        if salvar_necessario:
            wb.save(ARQUIVO_FINAL)
            print(f" -> Configuração salva com sucesso.")
        
        wb.close()
        
    except Exception as e:
        if "Permission denied" in str(e):
            print("ERRO CRÍTICO: O Excel está aberto. Feche-o.")
        logging.error(f"Erro ao inicializar config: {e}")

def carregar_configuracoes_do_excel():
    """
    Lê a aba 'config_fornecedor' e transforma no dicionário.
    Agora lê também a coluna 'Tipo_leitura'.
    """
    inicializar_aba_config()
    
    config_dict = {} 
    NOME_ABA_CONFIG = "config_fornecedor"

    try:
        if not ARQUIVO_FINAL.exists():
             # Se der erro de arquivo inexistente, usa memória
             raise Exception("Arquivo não encontrado para leitura.")

        # Lê a aba correta
        df_config = pd.read_excel(ARQUIVO_FINAL, sheet_name=NOME_ABA_CONFIG)
        df_config = df_config.dropna(how='all') 

        for _, row in df_config.iterrows():
            fornecedor = str(row["Fornecedor"])
            if fornecedor == 'nan': continue

            palavras = str(row["Palavras_Chave"]).split(',')
            palavras_limpas = [p.strip() for p in palavras if p.strip()]
            
            # --- NOVO: LÊ O TIPO DE LEITURA ---
            # Padronizamos para minúsculo para facilitar a comparação (corpo, pdf, ambos)
            tipo_leitura = str(row.get("Tipo_leitura", "corpo")).strip().lower()
            
            # Validação simples
            if tipo_leitura not in ['corpo', 'pdf', 'ambos']:
                tipo_leitura = 'corpo'
            # ----------------------------------

            config_dict[fornecedor] = {
                "assuntos_possiveis": palavras_limpas,
                "nome_aba": str(row["Nome_Aba"]),
                "classificacao_opex": str(row["Categoria_OPEX"]),
                "tipo_leitura": tipo_leitura, # Guardamos na config
                "colunas_renomear": {}
            }
        
        print(f" -> Configuração carregada: {len(config_dict)} fornecedores.")
        return config_dict

    except Exception as e:
        logging.error(f"Erro lendo Excel ({e}). Usando padrão de memória.")
        print(f"Aviso: Usando padrão de memória (Erro: {e})")
        
        # Fallback memória
        fallback_dict = {}
        dados = CONFIG_PADRAO
        if isinstance(CONFIG_PADRAO, dict): dados = list(CONFIG_PADRAO.values())

        for item in dados:
            if isinstance(item, dict):
                fallback_dict[item["Fornecedor"]] = {
                    "assuntos_possiveis": item["Palavras_Chave"].split(','),
                    "nome_aba": item["Nome_Aba"],
                    "classificacao_opex": item["Categoria_OPEX"],
                    "tipo_leitura": item.get("Tipo_leitura", "corpo").lower(),
                    "colunas_renomear": {}
                }
        return fallback_dict

def enviar_email_resumo(resumo_dados):
    """
    Gera o e-mail e MOSTRA NA TELA (.Display) para conferência.
    """

    if not resumo_dados:
        print(" -> Nada processado. E-mail de resumo não será gerado.")
        logging.info("Nada processado. E-mail ignorado.")
        return
    
    try:
        outlook = win32com.client.Dispatch("Outlook.Application")
        mail = outlook.CreateItem(0)

        # SEU E-MAIL AQUI
        mail.To = "matheus.augusto@grupofleury.com.br; lucas.rosa@grupofleury.com.br"
        mail.CC = "gustavo.guanas@grupofleury.com.br; jeferson.costa@grupofleury.com.br"

        data_hoje = datetime.now().strftime('%d/%m/%Y')
        mail.Subject = f"Resumo Processamento OPEX - {data_hoje}"

        #--- MONTAGEM DO HTML ---
        html = f"""
        <html>
        <body style="font-family: Calibri, sans-serif;">
            <h2 style="color: #2F75B5;">Processamento Finalizado com Sucesso</h2>
            <p>O robô Sentinela OPEX finalizou a execução de hoje ({data_hoje}).</p>
            
            <table style="border-collapse: collapse; width: 100%; max-width: 600px;">
                <tr style="background-color: #f2f2f2;">
                    <th style="border: 1px solid #ddd; padding: 8px; text-align: left;">Fornecedor</th>
                    <th style="border: 1px solid #ddd; padding: 8px; text-align: center;">Qtd E-mails</th>
                    <th style="border: 1px solid #ddd; padding: 8px; text-align: right;">Valor Total</th>
                </tr>
        """

        total_geral_valor = 0
        total_geral_qtd = 0

        for forn, dados in resumo_dados.items():
            qtd = dados.get('qtd', 0)
            valor = dados.get('valor', 0.0)
            total_geral_valor += valor
            total_geral_qtd += qtd

            valor_fmt = f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            
            html += f"""
                <tr>
                    <td style="border: 1px solid #ddd; padding: 8px;">{forn}</td>
                    <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{qtd}</td>
                    <td style="border: 1px solid #ddd; padding: 8px; text-align: right;">{valor_fmt}</td>
                </tr>
            """

        total_fmt = f"R$ {total_geral_valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        html += f"""
                <tr style="font-weight: bold; background-color: #e6f3ff;">
                    <td style="border: 1px solid #ddd; padding: 8px;">TOTAL GERAL</td>
                    <td style="border: 1px solid #ddd; padding: 8px; text-align: center;">{total_geral_qtd}</td>
                    <td style="border: 1px solid #ddd; padding: 8px; text-align: right;">{total_fmt}</td>
                </tr>
            </table>
            <br>
            <p style="font-size: 11px; color: #888;">Este e-mail foi gerado automaticamente pelo Robô Sentinela OPEX.</p>
        </body>
        </html>
        """

        mail.HTMLBody = html
        
        # --- MUDANÇA CRUCIAL AQUI ---
        # mail.Send()  <-- Comentado para evitar bloqueio silencioso
        mail.Display() # <-- VAI ABRIR A JANELA DO E-MAIL NA SUA TELA
        
        print("\n -> JANELA DE E-MAIL ABERTA! VERIFIQUE E CLIQUE EM ENVIAR.")
        logging.info("Janela de e-mail de resumo aberta.")

    except Exception as e:
        print(f"Erro ao gerar e-mail de resumo: {e}")
        logging.error(f"Erro no e-mail: {e}")



def executar_pipeline():
    print("\n--- INICIANDO PROCESSAMENTO (CORRIGIDO E OTIMIZADO) ---")
    
    if is_file_open(ARQUIVO_FINAL):
        print("ERRO CRÍTICO: O arquivo Excel está aberto. Por favor, feche-o.")
        return
    
    realizar_backup_seguranca()
    
    # Cria pasta temporária para PDFs
    PASTA_TEMP_PDF = CAMINHO_PASTA_LOCAL / "Temp_PDF"
    if not PASTA_TEMP_PDF.exists(): PASTA_TEMP_PDF.mkdir()

    stats_geral = {} 

    try:
        outlook = win32com.client.Dispatch("Outlook.Application").GetNamespace("MAPI")
        inbox = outlook.GetDefaultFolder(6)
        msgs_para_mover_com_destino = []

        print(" -> Carregando configurações...")
        config_atual = carregar_configuracoes_do_excel()

        if not config_atual:
            print("ERRO: Nenhuma configuração carregada.")
            return

        for fornecedor, config in config_atual.items():
            modo_leitura = config.get('tipo_leitura', 'corpo').upper()
            print(f"\nVerificando: {fornecedor} (Modo: {modo_leitura})")
            
            mensagens = inbox.Items
            mensagens.Sort("[ReceivedTime]", True)
            assuntos_alvo = [a.lower() for a in config.get("assuntos_possiveis", [])]
            dfs_novos = []

            for msg in mensagens:
                try:
                    if getattr(msg, 'Class', 0) != 43: continue
                    try:
                        if msg.ReceivedTime.replace(tzinfo=None) < DATA_INICIO_LEITURA: continue
                    except: continue

                    assunto_msg = msg.Subject.lower()
                    if not any(alvo in assunto_msg for alvo in assuntos_alvo): continue 
                    
                    remetente_email = obter_email_remetente(msg)
                    
                    # --- BLOQUEIO DE REMETENTE ---
                    if remetente_email in REMETENTES_IGNORAR: 
                        print(f" -> Ignorado: Remetente bloqueado ({remetente_email})")
                        continue
                    # -----------------------------

                    print(f" -> PROCESSANDO: {msg.Subject}")
                    tabelas_encontradas = []
                    tipo_leitura = config.get('tipo_leitura', 'corpo')

                    # 1. TENTA LER HTML (CORPO)
                    if tipo_leitura in ['corpo', 'ambos']:
                        try:
                            # Tenta ler com lxml (melhor parser), fallback para bs4 se necessário
                            html_tables = pd.read_html(StringIO(msg.HTMLBody), header=0, flavor=['lxml', 'bs4'])
                            tabelas_encontradas.extend(html_tables)
                        except: 
                            # Tenta sem header se falhar
                            try:
                                html_tables = pd.read_html(StringIO(msg.HTMLBody), header=None, flavor=['lxml', 'bs4'])
                                tabelas_encontradas.extend(html_tables)
                            except: pass
                    
                    # 2. TENTA LER PDF (ANEXO)
                    if tipo_leitura in ['pdf', 'ambos']:
                        for att in msg.Attachments:
                            if att.FileName.lower().endswith('.pdf'):
                                caminho_temp = PASTA_TEMP_PDF / att.FileName
                                try:
                                    att.SaveAsFile(str(caminho_temp))
                                    print(f"    Anexo baixado: {att.FileName}")
                                    tabelas_pdf = extrair_tabelas_de_pdf(caminho_temp) # Usa a versão robusta
                                    tabelas_encontradas.extend(tabelas_pdf)
                                except Exception as e_pdf:
                                    logging.error(f"Erro PDF: {e_pdf}")
                                finally:
                                    if caminho_temp.exists(): os.remove(caminho_temp)

                    if not tabelas_encontradas:
                        logging.warning("Nenhuma tabela encontrada.")
                        print("    [AVISO] Nenhuma tabela encontrada (HTML ou PDF).")
                        continue

                    # --- LÓGICA DE SELEÇÃO DA MELHOR TABELA ---
                    tabela_alvo = None
                    colunas_esperadas = list(config['colunas_renomear'].keys())
                    
                    # Tenta achar a tabela certa na lista de tabelas encontradas (usando a função melhorada)
                    for tb in tabelas_encontradas:
                        tb_ajustada = encontrar_cabecalho_correto(tb, colunas_esperadas)
                        if tb_ajustada is not None:
                            tabela_alvo = tb_ajustada
                            break
                    
                    # Fallback Genérico: Se não achou pelo cabeçalho, pega qualquer uma que tenha termos financeiros
                    if tabela_alvo is None:
                        for tb in tabelas_encontradas:
                            cols_texto = [str(c).lower() for c in tb.columns]
                            if any(t in c for c in cols_texto for t in ['valor', 'total', 'r$', 'liquido', 'bruto']):
                                tabela_alvo = tb
                                print("    [INFO] Tabela identificada por termos financeiros (Fallback).")
                                break

                    if tabela_alvo is None:
                        logging.warning(f"Estrutura irreconhecível em: {msg.Subject}")
                        print("    [AVISO] Tabelas encontradas, mas nenhuma com estrutura válida.")
                        continue
                    
                    # Extração de Dados Finais
                    mes_pt, mes_num, ano_full = extrair_data_da_tabela(tabela_alvo)
                    if not mes_pt: mes_pt, mes_num, ano_full = extrair_info_data_inteligente(msg.Subject, msg.Body)
                    
                    metadados = {
                        'data_recebimento': str(msg.ReceivedTime),
                        'remetente': msg.SenderName,
                        'assunto': msg.Subject,
                        'mes_nome_pt': mes_pt, 
                        'ano_full': ano_full   
                    }

                    df_tratado = tratar_dataframe(tabela_alvo, config, metadados)
                    
                    # Verificação final antes de adicionar
                    if not df_tratado.empty:
                        dfs_novos.append(df_tratado)
                        pasta_ano = f"Processados_OPEX {ano_full}"
                        pasta_mes = f"{mes_num} - {mes_pt}" if mes_pt else "00 - A Classificar"
                        msgs_para_mover_com_destino.append((msg, pasta_ano, pasta_mes, fornecedor))
                        print(f"    [SUCESSO] {len(df_tratado)} linhas extraídas.")
                    else:
                        print("    [AVISO] Tabela encontrada mas vazia após tratamento.")

                except Exception as e:
                    logging.error(f"Erro msg: {e}")
                    continue
            
            if dfs_novos:
                full_new_data = pd.concat(dfs_novos)
                salvar_com_append_preservando_formatacao(full_new_data, ARQUIVO_FINAL, config['nome_aba'])
                
                # Coleta Stats para o e-mail
                qtd = len(dfs_novos)
                soma_valor = 0.0
                for col in full_new_data.columns:
                    if eh_coluna_financeira(col):
                        soma_valor = full_new_data[col].sum()
                        break
                stats_geral[fornecedor] = {'qtd': qtd, 'valor': soma_valor}
            else:
                logging.info(f"Nenhum dado novo para {fornecedor}.")

        print("\nMovendo e-mails processados...")
        for msg, pasta_ano_nome, pasta_mes_nome, pasta_fornecedor_nome in msgs_para_mover_com_destino:
            try:
                msg.UnRead = False 
                try: pasta_ano = inbox.Folders(pasta_ano_nome)
                except: pasta_ano = inbox.Folders.Add(pasta_ano_nome)
                try: pasta_mes = pasta_ano.Folders(pasta_mes_nome)
                except: pasta_mes = pasta_ano.Folders.Add(pasta_mes_nome)
                try: pasta_final = pasta_mes.Folders(pasta_fornecedor_nome)
                except: pasta_final = pasta_mes.Folders.Add(pasta_fornecedor_nome)
                msg.Move(pasta_final)
            except: pass
        
        # Envio de E-mail (Agora envia direto, sem .Display)
        print("\nGerando relatório e enviando e-mail...")
        enviar_email_resumo(stats_geral)
        
        if PASTA_TEMP_PDF.exists(): shutil.rmtree(PASTA_TEMP_PDF)
        print("\n--- PROCESSO CONCLUÍDO COM SUCESSO ---")

    except Exception as e:
        logging.critical(f"Falha critica: {e}")
        print(f"Erro Crítico: {e}")

if __name__ == "__main__":
    executar_pipeline()